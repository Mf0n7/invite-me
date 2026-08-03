"""Parse de planilhas (xlsx/csv) para extrair nomes de convidados.

Limites de segurança aplicados aqui (o tamanho do arquivo já foi checado em
`apps.common.validators.validate_spreadsheet_upload`):

* `MAX_ROWS` — corta a leitura cedo. Um .xlsx é um ZIP: 2 MB comprimidos podem
  virar milhões de linhas descomprimidas (*zip bomb*) e estourar a memória do
  worker. Lemos em streaming e paramos no teto.
* Nomes truncados em 120 caracteres (mesmo limite do model).
"""
import csv
import io

from openpyxl import load_workbook

HEADER_WORDS = {"nome", "name", "convidado", "guest", "nome do convidado"}
MAX_ROWS = 5_000
MAX_NAME_LENGTH = 120


class ImportError_(ValueError):
    """Erro de importação com mensagem amigável."""


def parse_guest_names(uploaded_file) -> list[str]:
    filename = (getattr(uploaded_file, "name", "") or "").lower()
    if filename.endswith(".csv"):
        rows = _read_csv(uploaded_file)
    elif filename.endswith((".xlsx", ".xlsm")):
        rows = _read_xlsx(uploaded_file)
    else:
        raise ImportError_("Formato não suportado. Envie um arquivo .csv ou .xlsx.")
    return _clean(rows)


def _clean(values: list[str]) -> list[str]:
    names: list[str] = []
    for i, raw in enumerate(values):
        value = (raw or "").strip()
        if not value:
            continue
        if i == 0 and value.lower() in HEADER_WORDS:  # ignora cabeçalho
            continue
        names.append(value[:MAX_NAME_LENGTH])
    return names


def _read_csv(f) -> list[str]:
    data = f.read()
    if isinstance(data, bytes):
        data = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(data))
    rows: list[str] = []
    for row in reader:
        if len(rows) >= MAX_ROWS:
            raise ImportError_(f"Planilha muito grande. O limite é {MAX_ROWS} linhas.")
        rows.append(row[0] if row else "")
    return rows


def _read_xlsx(f) -> list[str]:
    try:
        wb = load_workbook(f, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — arquivo corrompido/forjado
        raise ImportError_("Não foi possível ler a planilha enviada.") from exc
    try:
        ws = wb.active
        rows: list[str] = []
        for row in ws.iter_rows(values_only=True):
            if len(rows) >= MAX_ROWS:
                raise ImportError_(f"Planilha muito grande. O limite é {MAX_ROWS} linhas.")
            rows.append(str(row[0]) if row and row[0] is not None else "")
        return rows
    finally:
        wb.close()
